
import openpyxl
import json

wb = openpyxl.load_workbook(r'C:/Users/user/Desktop/예산작업/나노센터 예산.xlsx', data_only=True)

result = {}
result['sheets'] = wb.sheetnames

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 80:
            break
        if any(v is not None for v in row):
            rows.append({'row': i+1, 'data': [str(v) if v is not None else None for v in row]})
    result[sname] = rows

with open(r'C:/Users/user/Desktop/예산작업/excel_structure.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print('done')
