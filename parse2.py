
import openpyxl
import json

wb = openpyxl.load_workbook(r"C:/Users/user/Desktop/예산작업/나노센터 예산.xlsx", data_only=True)

# 지출예산 시트 전체 읽기
ws = wb["지출예산"]
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 200:
        break
    if any(v is not None for v in row):
        rows.append({"row": i+1, "data": [str(v) if v is not None else None for v in row]})

with open(r"C:/Users/user/Desktop/예산작업/expense_sheet.json", "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

# 예산총괄표 읽기
ws2 = wb["3-1.예산총괄표(완료)"]
rows2 = []
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if any(v is not None for v in row):
        rows2.append({"row": i+1, "data": [str(v) if v is not None else None for v in row]})

with open(r"C:/Users/user/Desktop/예산작업/summary_sheet.json", "w", encoding="utf-8") as f:
    json.dump(rows2, f, ensure_ascii=False, indent=2)

# 성질별 지출예산 읽기
ws3 = wb["3-5.지출예산 성질별총괄표(완료)"]
rows3 = []
for i, row in enumerate(ws3.iter_rows(values_only=True)):
    if any(v is not None for v in row):
        rows3.append({"row": i+1, "data": [str(v) if v is not None else None for v in row]})

with open(r"C:/Users/user/Desktop/예산작업/expense_nature.json", "w", encoding="utf-8") as f:
    json.dump(rows3, f, ensure_ascii=False, indent=2)

print("done")
print(f"지출예산 행수: {ws.max_row}")
print(f"총괄표 행수: {ws2.max_row}")
print(f"성질별 행수: {ws3.max_row}")
