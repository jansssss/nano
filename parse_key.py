
import openpyxl
import json

wb = openpyxl.load_workbook(r"C:/Users/user/Desktop/예산작업/나노센터 예산.xlsx", data_only=True)

target_sheets = ["수입예산", "지출예산", "3-1.예산총괄표(완료)", "3-4. 수입예산 성질별 총괄표(완료)", "3-5.지출예산 성질별총괄표(완료)"]

result = {}

for sname in target_sheets:
    if sname not in wb.sheetnames:
        print(f"시트 없음: {sname}")
        continue
    ws = wb[sname]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 120:
            break
        if any(v is not None for v in row):
            rows.append({"row": i+1, "data": [str(v) if v is not None else None for v in row]})
    result[sname] = rows
    print(f"시트 {sname}: {ws.max_row}행 x {ws.max_column}열")

with open(r"C:/Users/user/Desktop/예산작업/key_sheets.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("done")
